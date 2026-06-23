import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

STATUS = {
    # Auth
    'User sign up with email/password': ('Working', 'AuthModal SignupForm + store.signup()'),
    'User login with email/password': ('Working', 'AuthModal LoginForm + store.login()'),
    'Google login support': ('Working', 'store.googleLogin() + Google OAuth button'),
    'Logout': ('Working', 'store.logout() wired in Navbar and ProfileModal'),
    'Profile creation in Supabase (profiles)': ('Working', 'DB trigger auto-creates profile on auth signup'),
    'Profile update (name; email; password)': ('Working', 'ProfileModal form + store.updateProfile()'),
    'Protected actions require login': ('Working', 'requireAuth() gates tickets / favorites / cart / profile'),
    # Homepage
    'Hero landing section': ('Working', 'Hero.jsx with heading, subtext, search bar, category pills, sort'),
    'Search events by keyword': ('Working', 'EventGrid useMemo filters title + desc by search string'),
    'Filter events by city': ('Working', 'Hero city dropdown -> EventGrid matchCity filter'),
    'Filter events by category': ('Working', 'Hero category pills -> EventGrid matchCat filter'),
    'Sort events by date': ('Working', 'Sort dropdown: date / price-asc / price-desc / name A-Z'),
    'Sticky responsive navbar': ('Working', 'Navbar.jsx present with responsive right-side controls'),
    'Clickable logo returns to homepage': ('Not Implemented', 'Logo img has no onClick or anchor wrapper'),
    'Custom favicon': ('Working', 'public/favicon.png linked via <link rel=icon> in index.html'),
    'Custom branded logo in navbar': ('Working', 'public/ouimoove-logo.png with text fallback'),
    # Events Discovery
    'Event listing grid': ('Working', 'EventGrid.jsx renders EventCard grid with useMemo filtering'),
    'Event detail modal': ('Working', 'EventDetailModal.jsx: image, tickets, qty controls, add to cart'),
    'Event category display': ('Working', 'Category badge on EventCard and in EventDetailModal'),
    'Event city/location display': ('Working', 'City shown on EventCard footer and in EventDetailModal meta'),
    'Event date/time display': ('Working', 'Date on card; date + time in detail modal'),
    'Ticket availability display': ('Working', 'EventDetailModal shows "avail = total - sold" per ticket type'),
    'Ticket pricing display': ('Working', 'Min price on EventCard; per-type price in EventDetailModal'),
    'Event descriptions': ('Working', 'Shown on EventCard (truncated) and EventDetailModal (full)'),
    'Default seeded events from database': ('Working', 'loadEvents() fetches Supabase; falls back to DEFAULT_EVENTS'),
    'Organizer-created events display live': ('Working', 'createEvent() calls loadEvents() after publish'),
    # Favorites
    'Add event to favorites': ('Working', 'store.toggleFavorite() inserts to Supabase favorites table'),
    'Remove event to favorites': ('Working', 'store.toggleFavorite() deletes from Supabase favorites table'),
    'Favorites saved in Supabase': ('Working', 'Supabase favorites table: user_id + event_id'),
    'Favorites modal view': ('Working', 'FavoritesModal.jsx with list, view and remove actions'),
    'Favorites require login': ('Working', 'requireAuth() gates onFavorites + handleToggleFav'),
    # Cart
    'Add tickets to cart': ('Working', 'EventDetailModal -> store.addToCart()'),
    'Multiple ticket type selection': ('Working', 'EventDetailModal shows each ticket type with +/- qty controls'),
    'Remove tickets from cart': ('Working', 'CartModal remove button -> store.removeFromCart()'),
    'Cart badge count in navbar': ('Working', 'Navbar shows cartCount badge on cart button'),
    'Cart total calculation': ('Working', 'cartTotal derived in store from cart items prices x qty'),
    'Checkout modal': ('Working', 'CheckoutModal.jsx: order summary + method selector + fields'),
    'Payment method selection (simulation)': ('Working', 'Card (inline fields) / T-Money / Flooz with phone input'),
    'Purchase confirmation flow': ('Working', 'handlePurchase -> store.purchase() -> success toast'),
    # Orders
    'Orders saved in Supabase (orders)': ('Working', 'store.purchase() inserts to orders table with status=paid'),
    'Order items saved in Supabase (order_items)': ('Working', 'store.purchase() inserts all cart items to order_items'),
    'Ticket stock updates after purchase': ('Working', 'purchase() updates ticket_types.quantity_sold'),
    '“Mes Billets” modal': ('Working', 'MyTicketsModal.jsx with TicketCard per order'),
    'Purchased tickets display': ('Working', 'TicketCard: event title, ticket types, qty, total, date'),
    'User ticket history': ('Working', 'loadMyOrders() fetches all orders for user ordered by date desc'),
    'Purchase persistence after refresh': ('Working', 'Orders fetched from Supabase on auth session restore'),
    # Organizer
    'Organizer dashboard modal': ('Working', 'OrganizerModal.jsx with tabs: overview / events / create / attendees / admin'),
    'Overview statistics': ('Working', 'OverviewTab stats grid from organizer_stats RPC'),
    'Revenue tracking': ('Working', 'organizerStats.total_revenue_cfa displayed'),
    'Tickets sold tracking': ('Working', 'organizerStats.total_tickets_sold displayed'),
    'Participant count': ('Working', 'organizerStats.unique_attendees displayed'),
    'Created events count': ('Working', 'organizerStats.published_events displayed'),
    # Event Management
    'Create new event': ('Working', 'CreateTab form + store.createEvent() publishes to Supabase'),
    'Delete created event': ('Working', 'EventsTab delete button + store.deleteEvent()'),
    'Multiple ticket types per event': ('Working', 'CreateTab: dynamic ticketTypes array with add/remove rows'),
    'Ticket quantity management': ('Working', 'qty field per ticket type in CreateTab'),
    'Ticket pricing setup': ('Working', 'price field per ticket type in CreateTab'),
    'Event category selection': ('Working', 'CATEGORIES dropdown in CreateTab'),
    'Event city selection': ('Working', 'CITIES dropdown in CreateTab'),
    'Event date/time selection': ('Working', 'Date + time inputs in CreateTab'),
    'Event description': ('Working', 'Textarea in CreateTab'),
    'Event emoji': ('Working', 'Emoji text input in CreateTab (defaults to ticket emoji)'),
    'Event publishing to Supabase': ('Working', 'createEvent() inserts with status=published'),
    # Stock
    'Live stock deduction after purchase': ('Working', 'purchase() updates quantity_sold after order confirmed'),
    'Sold/remaining ticket tracking': ('Working', 'EventDetailModal: avail = total - sold per ticket type'),
    'Prevent overselling logic': ('Partial', 'UI disables + button at avail limit; no server-side constraint'),
    # Attendees
    'Participants tab': ('Working', 'AttendeesTab in OrganizerModal'),
    'View attendees per event': ('Working', 'Dropdown filters attendees by event'),
    'Organizer attendee list': ('Working', 'loadOrganizerOrders() uses organizer_attendees DB view for real names'),
    'Check-in button': ('Working', 'Per-attendee check-in button in AttendeesTab'),
    'Real check-in persistence in Supabase': ('Working', 'checkinPurchase() updates order_items.checked_in + checked_in_at'),
    'Checked-in state display': ('Working', 'Green badge shown for checked-in attendees; count in stats bar'),
    # Images
    'Supabase Storage bucket created': ('Unverifiable', 'Cannot confirm without Supabase admin access'),
    'Image upload during event creation': ('Not Implemented', 'CreateTab only has URL text input; no file upload picker'),
    'Public image URL generation': ('Not Implemented', 'No upload capability; organizer must supply external URL'),
    'Image saved to events.image_url': ('Working', 'imageUrl from form stored in DB; shown on EventCard + detail modal'),
    # Deployment
    'Vercel production deployment': ('Working', 'Latest deploy READY on Vercel (commit fef3c34)'),
    'GitHub deployment flow': ('Working', 'Auto-deploys on push to main via Vercel Git integration'),
    'Supabase backend integration': ('Working', 'supabase.js client configured; all queries through Supabase'),
    'Supabase database tables': ('Working', 'orders, order_items, events, ticket_types, profiles, favorites, organizer_applications'),
    'Supabase RLS policies': ('Unverifiable', 'Cannot confirm from frontend code; assumed configured in Supabase dashboard'),
    'Supabase Storage integration': ('Unverifiable', 'No file upload UI; Storage bucket not actively used yet'),
    # Next features
    'Real payment gateway integration': ('Not Implemented', 'PayDunya redirect removed; now direct DB insert simulation only'),
    'Email confirmations': ('Not Implemented', 'No email sending code anywhere in the codebase'),
    'QR ticket validation': ('Working', 'MyTicketsModal generates QR via qrcode library keyed to order ID'),
    'Organizer/admin role permissions': ('Working', 'userRole + isOrganizer/isAdmin flags; Admin tab for promotions'),
    'Event editing': ('Not Implemented', 'Only create and delete; no edit form exists'),
    'Refund system': ('Not Implemented', 'Not implemented'),
    'Promo codes': ('Not Implemented', 'Not implemented'),
    'Analytics dashboard': ('Not Implemented', 'Basic stats in overview tab only; no dedicated analytics view'),
    'Push notifications': ('Not Implemented', 'Not implemented'),
    'Mobile app version': ('Not Implemented', 'Not implemented'),
}

FILL = {
    'Working':        'FF2ECC71',
    'Not Implemented':'FFFF6B6B',
    'Partial':        'FFFFC300',
    'Unverifiable':   'FFB0BEC5',
}
TEXT = {
    'Working':        'FF0A3622',
    'Not Implemented':'FF6B0000',
    'Partial':        'FF4A3800',
    'Unverifiable':   'FF2C3E50',
}

wb = openpyxl.load_workbook(r'C:\Users\Jose\Downloads\OuiMoove App Development Detailed Checklist.xlsx')
ws = wb.active

# Header row
for col, label in [(4, 'Status'), (5, 'Notes')]:
    c = ws.cell(row=1, column=col, value=label)
    c.font = Font(name='Arial', bold=True, color='FFFFFFFF')
    c.fill = PatternFill('solid', start_color='FF2C2C3E')
    c.alignment = Alignment(horizontal='center', vertical='center')

ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 60

for row in ws.iter_rows(min_row=2):
    feature = row[1].value
    if not feature:
        continue
    result = STATUS.get(str(feature).strip())
    if not result:
        continue
    status, note = result
    fill_color = FILL.get(status, 'FFEEEEEE')
    text_color = TEXT.get(status, 'FF000000')

    d = ws.cell(row=row[0].row, column=4, value=status)
    d.font = Font(name='Arial', bold=True, color=text_color)
    d.fill = PatternFill('solid', start_color=fill_color)
    d.alignment = Alignment(horizontal='center', vertical='center')

    e = ws.cell(row=row[0].row, column=5, value=note)
    e.font = Font(name='Arial', size=9)
    e.alignment = Alignment(vertical='center', wrap_text=True)

    ws.row_dimensions[row[0].row].height = 32

out = r'C:\Users\Jose\Desktop\ouimoove\OuiMoove Checklist - Audit Results.xlsx'
wb.save(out)
print('Saved:', out)

# Print summary
from collections import Counter
counts = Counter(v[0] for v in STATUS.values())
print('\nSummary:')
for k, n in sorted(counts.items()):
    print(f'  {k}: {n}')
print(f'  TOTAL: {sum(counts.values())}')
